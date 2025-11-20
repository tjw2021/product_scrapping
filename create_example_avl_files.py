"""
Create Example AVL Files
Generates sample THRIVE and GOODLEAP AVL files for testing and as templates
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


def create_thrive_avl():
    """Create example THRIVE AVL file"""
    print("Creating THRIVE AVL example file...")

    # Sample data - common solar manufacturers and models
    data = [
        # Solar Panels - Domestic Content
        {'Manufacturer': 'Silfab', 'Model': 'SIL-380-BX', 'Domestic Content': 'Yes'},
        {'Manufacturer': 'Silfab', 'Model': 'SIL-400-BX', 'Domestic Content': 'Yes'},
        {'Manufacturer': 'Qcells', 'Model': 'Q.PEAK DUO BLK ML-G10+ 400', 'Domestic Content': 'Yes'},
        {'Manufacturer': 'Mission Solar', 'Model': 'MSE375SQ8T', 'Domestic Content': 'Yes'},
        {'Manufacturer': 'Mission Solar', 'Model': 'MSE390SQ8T', 'Domestic Content': 'Yes'},

        # Solar Panels - Non-Domestic
        {'Manufacturer': 'Canadian Solar', 'Model': 'CS3W-400MS', 'Domestic Content': 'No'},
        {'Manufacturer': 'Canadian Solar', 'Model': 'CS3W-410MS', 'Domestic Content': 'No'},
        {'Manufacturer': 'JA Solar', 'Model': 'JAM72S30-540/MR', 'Domestic Content': 'No'},
        {'Manufacturer': 'Trina Solar', 'Model': 'TSM-DE09.08', 'Domestic Content': 'No'},
        {'Manufacturer': 'Longi', 'Model': 'LR5-72HIH-540M', 'Domestic Content': 'No'},

        # Inverters - Domestic Content
        {'Manufacturer': 'SolarEdge', 'Model': 'SE7600H-US', 'Domestic Content': 'Yes'},
        {'Manufacturer': 'SolarEdge', 'Model': 'SE10000H-US', 'Domestic Content': 'Yes'},
        {'Manufacturer': 'Enphase', 'Model': 'IQ8PLUS-72-2-US', 'Domestic Content': 'Yes'},
        {'Manufacturer': 'Enphase', 'Model': 'IQ8M-72-2-US', 'Domestic Content': 'Yes'},

        # Inverters - Non-Domestic
        {'Manufacturer': 'SMA', 'Model': 'SB7.7-1', 'Domestic Content': 'No'},
        {'Manufacturer': 'Fronius', 'Model': 'Primo 8.2-1', 'Domestic Content': 'No'},
        {'Manufacturer': 'Huawei', 'Model': 'SUN2000-10KTL-US', 'Domestic Content': 'No'},

        # Batteries - Domestic Content
        {'Manufacturer': 'Tesla', 'Model': 'Powerwall 3', 'Domestic Content': 'Yes'},
        {'Manufacturer': 'Enphase', 'Model': 'IQ Battery 5P', 'Domestic Content': 'Yes'},

        # Batteries - Non-Domestic
        {'Manufacturer': 'LG Energy Solution', 'Model': 'RESU10H', 'Domestic Content': 'No'},
        {'Manufacturer': 'BYD', 'Model': 'Battery-Box Premium HVS', 'Domestic Content': 'No'},
    ]

    df = pd.DataFrame(data)

    # Save to Excel
    filename = 'THRIVE_AVL.xlsx'
    df.to_excel(filename, index=False, sheet_name='THRIVE AVL')

    # Apply formatting
    wb = load_workbook(filename)
    ws = wb['THRIVE AVL']

    # Format header
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column_letter].width = max_length + 2

    wb.save(filename)
    print(f"✅ Created {filename} with {len(df)} entries")


def create_goodleap_avl():
    """Create example GOODLEAP AVL file"""
    print("Creating GOODLEAP AVL example file...")

    # Sample data - manufacturer-level approval (no specific models)
    data = [
        # Solar Panels - Domestic Content
        {'Manufacturer': 'Silfab', 'Program Type': 'Loans/Leases/PPAs', 'Domestic Content': 'Yes'},
        {'Manufacturer': 'Qcells', 'Program Type': 'Loans/Leases/PPAs', 'Domestic Content': 'Yes'},
        {'Manufacturer': 'Mission Solar', 'Program Type': 'Loans/Leases/PPAs', 'Domestic Content': 'Yes'},
        {'Manufacturer': 'Hanwha Q CELLS', 'Program Type': 'Loans/Leases/PPAs', 'Domestic Content': 'Yes'},

        # Solar Panels - Non-Domestic
        {'Manufacturer': 'Canadian Solar', 'Program Type': 'Loans/Leases/PPAs', 'Domestic Content': 'No'},
        {'Manufacturer': 'JA Solar', 'Program Type': 'Loans/Leases/PPAs', 'Domestic Content': 'No'},
        {'Manufacturer': 'Trina Solar', 'Program Type': 'Loans/Leases/PPAs', 'Domestic Content': 'No'},
        {'Manufacturer': 'Longi', 'Program Type': 'Loans/Leases/PPAs', 'Domestic Content': 'No'},
        {'Manufacturer': 'Jinko Solar', 'Program Type': 'Loans/Leases/PPAs', 'Domestic Content': 'No'},

        # Inverters - Domestic Content
        {'Manufacturer': 'SolarEdge', 'Program Type': 'Loans/Leases/PPAs', 'Domestic Content': 'Yes'},
        {'Manufacturer': 'Enphase', 'Program Type': 'Loans/Leases/PPAs', 'Domestic Content': 'Yes'},
        {'Manufacturer': 'Tesla', 'Program Type': 'Loans/Leases/PPAs', 'Domestic Content': 'Yes'},

        # Inverters - Non-Domestic
        {'Manufacturer': 'SMA', 'Program Type': 'Loans/Leases/PPAs', 'Domestic Content': 'No'},
        {'Manufacturer': 'Fronius', 'Program Type': 'Loans/Leases/PPAs', 'Domestic Content': 'No'},
        {'Manufacturer': 'Huawei', 'Program Type': 'Loans Only', 'Domestic Content': 'No'},
        {'Manufacturer': 'GoodWe', 'Program Type': 'Loans Only', 'Domestic Content': 'No'},
        {'Manufacturer': 'Solis', 'Program Type': 'Loans Only', 'Domestic Content': 'No'},

        # Batteries
        {'Manufacturer': 'LG Energy Solution', 'Program Type': 'Loans/Leases/PPAs', 'Domestic Content': 'No'},
        {'Manufacturer': 'BYD', 'Program Type': 'Loans Only', 'Domestic Content': 'No'},
        {'Manufacturer': 'Generac', 'Program Type': 'Loans/Leases/PPAs', 'Domestic Content': 'Yes'},
    ]

    df = pd.DataFrame(data)

    # Save to Excel
    filename = 'GOODLEAP_AVL.xlsx'
    df.to_excel(filename, index=False, sheet_name='GOODLEAP AVL')

    # Apply formatting
    wb = load_workbook(filename)
    ws = wb['GOODLEAP AVL']

    # Format header
    header_fill = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column_letter].width = max_length + 2

    wb.save(filename)
    print(f"✅ Created {filename} with {len(df)} entries")


def main():
    """Create both AVL files"""
    print("\n" + "="*60)
    print("Creating Example AVL Files")
    print("="*60 + "\n")

    create_thrive_avl()
    print()
    create_goodleap_avl()

    print("\n" + "="*60)
    print("✅ AVL Files Created Successfully!")
    print("="*60)
    print("\nFiles created:")
    print("  • THRIVE_AVL.xlsx")
    print("  • GOODLEAP_AVL.xlsx")
    print("\nThese files contain sample data. Replace with your actual AVL data.")
    print("="*60 + "\n")


if __name__ == "__main__":
    main()
