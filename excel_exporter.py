"""
Excel Exporter - Exports product data to organized Excel workbooks
Creates separate sheets for each equipment category with formatting
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import List, Dict, Optional
import os


class ExcelExporter:
    """Export data to organized Excel workbook with separate sheets by category"""

    # Define categories and their order
    CATEGORIES = [
        'Solar Panel',
        'Inverter',
        'Battery/Storage',
        'Charge Controller',
        'Transformer',
        'Switch',
        'Racking/Mounting',
        'BOS/Electrical',
        'Other'
    ]

    def __init__(self, output_filename: Optional[str] = None):
        """
        Initialize Excel exporter

        Args:
            output_filename: Output filename (will be auto-generated if not provided)
        """
        if output_filename is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_filename = f'solar_equipment_database_{timestamp}.xlsx'

        self.output_filename = output_filename
        self.writer = None

    def _sanitize_sheet_name(self, name: str) -> str:
        """
        Sanitize sheet name to meet Excel requirements

        Args:
            name: Original sheet name

        Returns:
            Sanitized sheet name
        """
        # Excel doesn't allow these characters in sheet names
        invalid_chars = ['\\', '/', '*', '?', ':', '[', ']']
        sanitized = name
        for char in invalid_chars:
            sanitized = sanitized.replace(char, '-')

        # Limit to 31 characters
        sanitized = sanitized[:31]

        return sanitized

    def export_by_category(
        self,
        products_df: pd.DataFrame,
        include_summary: bool = True,
        include_domestic_only: bool = True
    ):
        """
        Create separate sheets for each category

        Args:
            products_df: Product dataframe with all products
            include_summary: Whether to include summary sheet
            include_domestic_only: Whether to include domestic content only sheet
        """
        print(f"\n📊 Exporting to Excel: {self.output_filename}")

        # Ensure category column exists
        if 'category' not in products_df.columns:
            print("⚠️  'category' column not found, skipping category-based export")
            return

        # Create Excel writer
        self.writer = pd.ExcelWriter(self.output_filename, engine='openpyxl')

        # Track statistics for summary
        category_stats = {}

        # Create sheet for each category
        for category in self.CATEGORIES:
            category_df = products_df[products_df['category'] == category].copy()

            if category_df.empty:
                continue

            print(f"  Creating sheet: {category} ({len(category_df)} products)")

            # Prepare dataframe for export
            export_df = self._prepare_export_dataframe(category_df)

            # Create sheet name (sanitize and limit length)
            sheet_name = self._sanitize_sheet_name(category)

            # Write to sheet
            export_df.to_excel(self.writer, sheet_name=sheet_name, index=False)

            # Collect statistics
            category_stats[category] = self._calculate_category_stats(category_df)

        # Create summary sheet
        if include_summary and category_stats:
            print(f"  Creating sheet: Summary")
            self._create_summary_sheet(category_stats, products_df)

        # Create domestic content only sheet
        if include_domestic_only and 'domestic_content_qualified' in products_df.columns:
            domestic_df = products_df[products_df['domestic_content_qualified'] == True].copy()
            if not domestic_df.empty:
                print(f"  Creating sheet: Domestic Content ({len(domestic_df)} products)")
                export_df = self._prepare_export_dataframe(domestic_df)
                export_df.to_excel(self.writer, sheet_name='Domestic Content', index=False)

        # Create "All Products" sheet
        print(f"  Creating sheet: All Products ({len(products_df)} products)")
        export_df = self._prepare_export_dataframe(products_df)
        export_df.to_excel(self.writer, sheet_name='All Products', index=False)

        # Save workbook
        self.writer.close()

        # Apply formatting
        self._apply_formatting()

        print(f"\n✅ Excel export complete: {self.output_filename}")

    def _prepare_export_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Prepare dataframe for export by reordering and formatting columns

        Args:
            df: Source dataframe

        Returns:
            Formatted dataframe ready for export
        """
        # Define column order and labels
        column_order = [
            ('distributor', 'Distributor'),
            ('category', 'Category'),
            ('brand', 'Manufacturer'),
            ('sku', 'Model/SKU'),
            ('title', 'Product Title'),
            ('wattage', 'Wattage'),
            ('price_per_unit', 'Price Per Unit'),
            ('price', 'Total Price'),
            ('quantity', 'Quantity'),
            ('stock_status', 'Stock Status'),
            ('inventory_qty', 'Inventory Qty'),
            ('product_url', 'Product URL'),
            ('thrive_approved', 'THRIVE Approved'),
            ('thrive_domestic', 'THRIVE Domestic'),
            ('goodleap_approved', 'GOODLEAP Approved'),
            ('goodleap_domestic', 'GOODLEAP Domestic'),
            ('goodleap_program', 'GOODLEAP Program'),
            ('on_any_avl', 'On Any AVL'),
            ('domestic_content_qualified', 'Domestic Content'),
            ('last_updated', 'Last Updated')
        ]

        # Create export dataframe with only available columns
        export_data = {}
        for col_name, col_label in column_order:
            if col_name in df.columns:
                export_data[col_label] = df[col_name]

        export_df = pd.DataFrame(export_data)

        # Format boolean columns
        bool_columns = [
            'THRIVE Approved', 'THRIVE Domestic',
            'GOODLEAP Approved', 'GOODLEAP Domestic',
            'On Any AVL', 'Domestic Content'
        ]
        for col in bool_columns:
            if col in export_df.columns:
                export_df[col] = export_df[col].map({True: 'Yes', False: 'No', None: 'N/A'})

        # Format price columns
        price_columns = ['Price Per Unit', 'Total Price']
        for col in price_columns:
            if col in export_df.columns:
                export_df[col] = export_df[col].apply(
                    lambda x: f"${x:.2f}" if isinstance(x, (int, float)) and x > 0 else 'N/A'
                )

        return export_df

    def _calculate_category_stats(self, category_df: pd.DataFrame) -> Dict:
        """
        Calculate statistics for a category

        Args:
            category_df: Dataframe for the category

        Returns:
            Dictionary with statistics
        """
        stats = {
            'total': len(category_df),
            'in_stock': 0,
            'avg_price': 0,
            'min_price': 0,
            'max_price': 0,
            'thrive_approved': 0,
            'goodleap_approved': 0,
            'domestic_content': 0
        }

        # Stock status
        if 'stock_status' in category_df.columns:
            stats['in_stock'] = len(category_df[category_df['stock_status'] == 'In Stock'])

        # Price statistics
        if 'price_per_unit' in category_df.columns:
            prices = category_df['price_per_unit']
            prices = prices[prices > 0]  # Filter out zero/invalid prices
            if not prices.empty:
                stats['avg_price'] = prices.mean()
                stats['min_price'] = prices.min()
                stats['max_price'] = prices.max()

        # AVL statistics
        if 'thrive_approved' in category_df.columns:
            stats['thrive_approved'] = category_df['thrive_approved'].sum()

        if 'goodleap_approved' in category_df.columns:
            stats['goodleap_approved'] = category_df['goodleap_approved'].sum()

        if 'domestic_content_qualified' in category_df.columns:
            stats['domestic_content'] = category_df['domestic_content_qualified'].sum()

        return stats

    def _create_summary_sheet(self, category_stats: Dict, full_df: pd.DataFrame):
        """
        Create summary statistics sheet

        Args:
            category_stats: Dictionary of statistics per category
            full_df: Full product dataframe
        """
        summary_data = []

        for category in self.CATEGORIES:
            if category not in category_stats:
                continue

            stats = category_stats[category]

            summary_data.append({
                'Category': category,
                'Total Products': stats['total'],
                'In Stock': stats['in_stock'],
                'Avg Price': f"${stats['avg_price']:.2f}" if stats['avg_price'] > 0 else 'N/A',
                'Min Price': f"${stats['min_price']:.2f}" if stats['min_price'] > 0 else 'N/A',
                'Max Price': f"${stats['max_price']:.2f}" if stats['max_price'] > 0 else 'N/A',
                'THRIVE Approved': stats['thrive_approved'],
                'GOODLEAP Approved': stats['goodleap_approved'],
                'Domestic Content': stats['domestic_content']
            })

        # Add overall totals
        summary_data.append({
            'Category': 'TOTAL',
            'Total Products': len(full_df),
            'In Stock': full_df['stock_status'].eq('In Stock').sum() if 'stock_status' in full_df.columns else 0,
            'Avg Price': f"${full_df['price_per_unit'].mean():.2f}" if 'price_per_unit' in full_df.columns else 'N/A',
            'Min Price': f"${full_df['price_per_unit'].min():.2f}" if 'price_per_unit' in full_df.columns else 'N/A',
            'Max Price': f"${full_df['price_per_unit'].max():.2f}" if 'price_per_unit' in full_df.columns else 'N/A',
            'THRIVE Approved': full_df['thrive_approved'].sum() if 'thrive_approved' in full_df.columns else 0,
            'GOODLEAP Approved': full_df['goodleap_approved'].sum() if 'goodleap_approved' in full_df.columns else 0,
            'Domestic Content': full_df['domestic_content_qualified'].sum() if 'domestic_content_qualified' in full_df.columns else 0
        })

        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(self.writer, sheet_name='Summary', index=False)

    def _apply_formatting(self):
        """Apply formatting to all sheets in the workbook"""
        print("  Applying formatting...")

        # Load workbook
        wb = load_workbook(self.output_filename)

        # Define styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        total_row_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
        total_row_font = Font(bold=True, size=11)

        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Format each sheet
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Format header row
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)

                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass

                # Set width with min and max limits
                adjusted_width = min(max(max_length + 2, 10), 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Format Summary sheet TOTAL row
            if sheet_name == 'Summary':
                for row in ws.iter_rows(min_row=2):
                    if row[0].value == 'TOTAL':
                        for cell in row:
                            cell.fill = total_row_fill
                            cell.font = total_row_font

            # Freeze header row
            ws.freeze_panes = 'A2'

        # Save formatted workbook
        wb.save(self.output_filename)

    def export_simple(self, products_df: pd.DataFrame, sheet_name: str = 'Products'):
        """
        Simple export without category separation

        Args:
            products_df: Product dataframe
            sheet_name: Name for the sheet
        """
        print(f"\n📊 Exporting to Excel: {self.output_filename}")

        # Prepare dataframe
        export_df = self._prepare_export_dataframe(products_df)

        # Create Excel writer and save
        with pd.ExcelWriter(self.output_filename, engine='openpyxl') as writer:
            export_df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Apply formatting
        self._apply_formatting()

        print(f"✅ Excel export complete: {self.output_filename}")


if __name__ == "__main__":
    # Test the Excel exporter
    print("Testing Excel Exporter...")

    # Create sample data
    sample_data = [
        {
            'distributor': 'Solar Cellz USA',
            'category': 'Solar Panel',
            'brand': 'Canadian Solar',
            'sku': 'CS3W-400MS',
            'title': 'Canadian Solar 400W Panel',
            'wattage': '400W',
            'price': 150.00,
            'price_per_unit': 150.00,
            'quantity': 1,
            'stock_status': 'In Stock',
            'inventory_qty': '50',
            'product_url': 'https://example.com/product1',
            'thrive_approved': True,
            'thrive_domestic': False,
            'goodleap_approved': True,
            'goodleap_domestic': False,
            'goodleap_program': 'Loans/Leases/PPAs',
            'on_any_avl': True,
            'domestic_content_qualified': False,
            'last_updated': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        },
        {
            'distributor': 'Soligent',
            'category': 'Inverter',
            'brand': 'SMA',
            'sku': 'SB7.7-1',
            'title': 'SMA Sunny Boy 7.7kW Inverter',
            'wattage': 'N/A',
            'price': 2500.00,
            'price_per_unit': 2500.00,
            'quantity': 1,
            'stock_status': 'In Stock',
            'inventory_qty': '20',
            'product_url': 'https://example.com/product2',
            'thrive_approved': True,
            'thrive_domestic': True,
            'goodleap_approved': True,
            'goodleap_domestic': True,
            'goodleap_program': 'Loans Only',
            'on_any_avl': True,
            'domestic_content_qualified': True,
            'last_updated': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
    ]

    df = pd.DataFrame(sample_data)

    # Test export
    exporter = ExcelExporter('test_export.xlsx')
    exporter.export_by_category(df)

    print(f"\n✅ Test complete! Check test_export.xlsx")
